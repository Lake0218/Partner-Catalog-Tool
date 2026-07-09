from io import BytesIO
from inspect import Parameter, signature

import pandas as pd
from openpyxl import load_workbook

UPC_COL = "A"
OUTPUT_COL = "I"
START_ROW = 3
HEADER_ROWS = range(1, START_ROW)
HEADERS_TO_REMOVE = {"status", "errors", "warnings"}


def normalize_upc(value):
    if value is None or pd.isna(value):
        return None

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    return str(value).strip()


def normalize_header(value):
    if value is None or pd.isna(value):
        return ""

    return str(value).strip().lower()


def remove_output_columns(ws):
    columns_to_remove = []

    for column in range(1, ws.max_column + 1):
        headers = [
            normalize_header(ws.cell(row=row, column=column).value)
            for row in HEADER_ROWS
        ]
        if any(header in HEADERS_TO_REMOVE for header in headers):
            columns_to_remove.append(column)

    for column in reversed(columns_to_remove):
        ws.delete_cols(column)


def supports_query_ttl(query):
    try:
        parameters = signature(query).parameters
    except (TypeError, ValueError):
        return False

    return "ttl" in parameters or any(
        parameter.kind == Parameter.VAR_KEYWORD
        for parameter in parameters.values()
    )


def supports_query_params(query):
    try:
        parameters = signature(query).parameters
    except (TypeError, ValueError):
        return False

    return "params" in parameters or any(
        parameter.kind == Parameter.VAR_KEYWORD
        for parameter in parameters.values()
    )


def run_snowflake_query(conn, sql, params=None):
    if hasattr(conn, "query"):
        query_kwargs = {}
        if supports_query_ttl(conn.query):
            query_kwargs["ttl"] = 0
        if params is not None:
            if not supports_query_params(conn.query):
                raise ValueError(
                    "This Snowflake connection does not support query parameters."
                )
            query_kwargs["params"] = params
        return conn.query(sql, **query_kwargs)

    cur = conn.cursor()
    try:
        if params is None:
            cur.execute(sql)
        else:
            cur.execute(sql, params)
        rows = cur.fetchall()
        columns = [column[0] for column in cur.description]
    finally:
        cur.close()

    return pd.DataFrame(rows, columns=columns)


def load_sales_lookup(conn, business_id):
    if not business_id:
        raise ValueError("A business must be selected before querying Snowflake sales data.")

    sql = """
        WITH partner_barcodes AS (
            SELECT DISTINCT fc.barcode
            FROM fetch_services_prod.staging.fido_catalog_cleaned_stage fc
                LEFT JOIN analytics_prod.durin_service.p_brands b
                    ON fc.brand_id = b.id
                LEFT JOIN analytics_prod.durin_service.p_business biz
                    ON b.business_id = biz.id
            WHERE biz.id = %s
                AND fc.deleted_date IS NULL
                AND fc.added_date < DATEADD(MONTH, -6, CURRENT_DATE)
                AND COALESCE(b.deleted_ts, b.be_soft_deleted_ts) IS NULL
                AND biz.deleted_ts IS NULL
        )
        SELECT
            pb.barcode AS UPC,
            SUM(ri.final_sale) AS TOTAL_SALES
        FROM partner_barcodes AS pb
            LEFT JOIN analytics_prod.partner_analytics.pa_receipts_items AS ri
                ON ri.barcode = pb.barcode
                AND ri.purchase_date::DATE >= DATEADD(MONTH, -24, CURRENT_DATE)
        GROUP BY pb.barcode
        HAVING TOTAL_SALES IS NULL
        ORDER BY pb.barcode
    """

    df = run_snowflake_query(conn, sql, params=(business_id,))
    if df.empty:
        return {}

    df.columns = [str(column).upper() for column in df.columns]
    df["UPC"] = df["UPC"].apply(normalize_upc)
    df["TOTAL_SALES"] = pd.to_numeric(df["TOTAL_SALES"], errors="coerce").fillna(0)

    return dict(zip(df["UPC"], df["TOTAL_SALES"]))


def update_partner_catalog(workbook_file, sales_lookup):
    try:
        workbook_file.seek(0)
        wb = load_workbook(workbook_file)
    except Exception as exc:
        raise ValueError(
            "Could not load the uploaded workbook. Make sure it is a valid .xlsx file."
        ) from exc

    ws = wb[wb.sheetnames[0]]

    total_rows = 0
    zero_sales_rows = 0
    missing_rows = 0
    matched_rows = 0
    missing_upcs = []

    for row in range(START_ROW, ws.max_row + 1):
        upc = normalize_upc(ws[f"{UPC_COL}{row}"].value)
        if not upc:
            continue

        total_rows += 1
        has_zero_sales = upc in sales_lookup

        if has_zero_sales:
            matched_rows += 1
            ws[f"{OUTPUT_COL}{row}"] = "0 USD"
            zero_sales_rows += 1
        else:
            missing_rows += 1
            missing_upcs.append(upc)

    remove_output_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    summary = {
        "total_rows": total_rows,
        "zero_sales_rows": zero_sales_rows,
        "missing_rows": missing_rows,
        "missing_upcs": missing_upcs,
        "matched_rows": matched_rows,
    }

    return output, summary
